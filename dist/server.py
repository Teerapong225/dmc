from flask import Flask, render_template, request, redirect, flash, jsonify
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import shutil
import datetime
from flask_cors import CORS
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'my_secret_key_1234'  # ใช้สำหรับ flash message
CORS(app)

# โหลดไฟล์ Excel
file_path = "C:/Users/acer/Desktop/startbootstrap-heroic-features-master/dist/data/FoodAnalysis/analyzed_output.xlsx"
df = pd.read_excel(file_path, sheet_name="Sheets1",
                   names=["Business Date", "Country", "Major Group", "Menu Item Name", "Sales Count"])

# แปลงประเภทข้อมูลให้ถูกต้อง
df["Business Date"] = pd.to_datetime(df["Business Date"], errors="coerce")
df["Major Group"] = df["Major Group"].fillna("Unknown")  # แทนค่า NaN
df["Sales Count"] = pd.to_numeric(df["Sales Count"], errors="coerce").fillna(
    0)  # แปลง Count เป็นตัวเลข

# กำหนดโฟลเดอร์ที่เก็บไฟล์
UPLOAD_FOLDER_FOOD = 'C:/Users/acer/Desktop/startbootstrap-heroic-features-master/dist/data/Food'
UPLOAD_FOLDER_COUNTRY = 'C:/Users/acer/Desktop/startbootstrap-heroic-features-master/dist/data/Country'
MERGE_FOLDER = 'C:/Users/acer/Desktop/startbootstrap-heroic-features-master/dist/data/FoodAnalysis'

ALLOWED_EXTENSIONS = {'xlsx', 'csv'}

app.config['UPLOAD_FOLDER_FOOD'] = UPLOAD_FOLDER_FOOD
app.config['UPLOAD_FOLDER_COUNTRY'] = UPLOAD_FOLDER_COUNTRY
app.config['MERGE_FOLDER'] = MERGE_FOLDER

food_to_check = ['Business Date', 'Revenue Center', 'Check #', 'Open Time', 'Close Time', 'Menu Item #', 'Menu Item Name',
                 'Major Group', 'Family Group', 'Unit Price', 'Sales Count', 'Sales', 'Tender']  # ตัวอย่างคอลัมน์ที่ใช้สำหรับ Food

country_to_check = ['Room', 'Country', 'Arrival Date', 'Departure Date',
                    'Travel Agent Name', 'Folio Number', 'Check Number']  # ตัวอย่างคอลัมน์ที่ใช้สำหรับ Country


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ฟังก์ชันลดขนาดของ Check Number (หรือ Check #)
def reduce_check_number(data, column_name):
    print("ลดขนาดของ Check Number...")
    if column_name in data.columns:
        data[column_name] = data[column_name].astype(str).str[:8]
    print(f"เสร็จสิ้นการลดขนาดของ {column_name}")
    return data


# ฟังก์ชัน Fill Down (เติมค่าช่องว่างจากค่าด้านบน)
def fill_down_missing_values(data):
    print("กำลังเติมช่องว่าง...")
    return data.ffill()


# ฟังก์ชันเปลี่ยนชื่อแผ่นงานเป็น "Sheets1"
def rename_sheet_to_sheets1(file_path):
    print(f"กำลังเปลี่ยนชื่อแผ่นงานใน {file_path} เป็น Sheets1...")
    workbook = load_workbook(file_path)
    if "Sheet1" in workbook.sheetnames:
        sheet = workbook["Sheet1"]
    else:
        sheet = workbook.active
    sheet.title = "Sheets1"
    workbook.save(file_path)
    print(f"เปลี่ยนชื่อแผ่นงานเสร็จสิ้นใน {file_path}")


def add_table_borders(sheet):
    print("กำลังเพิ่มขอบตาราง...")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
    print("เพิ่มขอบตารางเสร็จสิ้น!")


def remove_empty_columns(file_path):
    df = pd.read_excel(file_path)
    to_remove = []
    for col in df.columns:
        # ลบเฉพาะคอลัมน์ที่ว่างเปล่าทั้งหมด
        if df[col].isna().all():
            to_remove.append(col)
    if to_remove:
        df.drop(columns=to_remove, inplace=True)
        df.to_excel(file_path, index=False)
        # แปลงชื่อ 'Unnamed: X' เป็น Excel letter
        converted = []
        for name in to_remove:
            if name.startswith("Unnamed: "):
                col_index = int(name.replace("Unnamed: ", ""))
                converted.append(get_column_letter(col_index + 1))
            else:
                converted.append(name)
        print(f"คอลัมน์ว่างคือ: {converted}")
        print("ลบคอลัมน์ว่างแล้ว")
    else:
        print("ไม่มีคอลัมน์ว่างให้ลบ")
    return df


# ฟังก์ชันปรับขนาดความกว้างคอลัมน์ให้พอดีกับข้อมูล
def adjust_column_width(data, file_path):
    print(f"กำลังปรับขนาดคอลัมน์ใน {file_path}...")
    workbook = load_workbook(file_path)
    sheet = workbook["Sheets1"]
    for col in sheet.columns:
        max_length = 0
        min_length = float('inf')
        column = col[0].column_letter  # Get the column name
        for cell in col:
            column = col[0].column_letter
            try:
                cell_length = len(
                    str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
                if cell_length < min_length:
                    min_length = cell_length
            except:
                pass
        adjusted_width = max(min_length + 2, min(max_length + 2, 30))
        sheet.column_dimensions[column].width = adjusted_width
    add_table_borders(sheet)  # เพิ่มขอบตาราง
    workbook.save(file_path)
    print(f"ปรับขนาดคอลัมน์เสร็จสิ้นใน {file_path}")


def append_timestamp_filename(upload_folder, filename):
    print(f"กำลังเพิ่ม timestamp ให้กับชื่อไฟล์ {filename}...")
    base, ext = os.path.splitext(filename)
    now = datetime.datetime.now().strftime("%y.%m.%d_%H.%M.%S")
    new_filename = f"{base}_{now}{ext}"
    print(f"ชื่อไฟล์ใหม่คือ {new_filename}")
    return new_filename


def remove_duplicate_rows(data, columns_to_check=None):
    print("กำลังลบแถวที่ซ้ำกัน...")
    try:
        if columns_to_check is None:
            initial_row_count = len(data)
            data = data.loc[~data.apply(
                lambda row: row.nunique() == 1, axis=1)]
            final_row_count = len(data)
            print(
                f"ลบแถวซ้ำ {initial_row_count - final_row_count} แถว (จากทุกคอลัมน์).")
        else:
            initial_row_count = len(data)
            data = data.loc[~data[columns_to_check].duplicated(keep='first')]
            final_row_count = len(data)
            print(
                f"ลบแถวซ้ำ {initial_row_count - final_row_count} แถว (จากคอลัมน์: {', '.join(columns_to_check)}).")
        return data
    except Exception as e:
        print(f"เกิดข้อผิดพลาด: {e}")
        return data


def process_uploaded_file(file, upload_folder, file_type):
    print(f"กำลังประมวลผลไฟล์ {file.filename}...")
    filename = file.filename
    file_ext = filename.rsplit('.', 1)[1].lower()
    filename = append_timestamp_filename(upload_folder, filename)
    file_path = os.path.join(upload_folder, filename)

    file.save(file_path)

    if file_ext == 'csv':
        try:
            print("กำลังอ่านไฟล์ CSV")
            data = pd.read_csv(file_path, encoding='utf-8',
                               delimiter=',', skiprows=5, on_bad_lines='skip')
        except pd.errors.ParserError:
            print("พบปัญหากับการอ่านไฟล์ CSV")
            return None

        file_path_xlsx = file_path.rsplit('.', 1)[0] + '.xlsx'
        print("กำลังเปลี่ยนเป็น xlsx")
        os.remove(file_path)  # ลบไฟล์ CSV เก่า
        print("กำลังลบไฟล์ csv เก่า")
        file_path = file_path_xlsx  # อัปเดต path เป็นไฟล์ใหม่
        print("กำลังบันทึกไฟล์ xlsx")
        data.to_excel(file_path, index=False, engine='openpyxl')
    print("กำลังอ่านไฟล์ xlsx")
    data = pd.read_excel(file_path)

    primary_keys = {
        "food": "Check #",
        "country": "Check Number"
    }
    required_key = primary_keys.get(file_type)
    check_row = data.apply(lambda row: row.astype(
        str).str.contains(required_key).any(), axis=1)

    if check_row.any():
        header_row_index = check_row.idxmax()
        data.columns = data.iloc[header_row_index]
        data = data.iloc[header_row_index + 1:]
        data = data.reset_index(drop=True)

    if file_type == "country":
        print("พบว่าเป็นไฟล์ country")
        data = reduce_check_number(data, "Check Number")
        data = fill_down_missing_values(data)
        columns_to_check = country_to_check
    elif file_type == "food":
        print("พบว่าเป็นไฟล์ food")
        data = data.iloc[:-3]
        columns_to_check = food_to_check

    if required_key not in data.columns:
        os.remove(file_path)  # ❌ ลบไฟล์ถ้าไม่มี Primary Key
        flash(f"This is not a valid {file_type} file.", "error")
        print(f"Error: This is not a valid {file_type} file.")
        return None

    data = remove_duplicate_rows(data, columns_to_check)
    print("กำลังบันทึกไฟล์")
    data.to_excel(file_path, index=False)

    remove_empty_columns(file_path)
    rename_sheet_to_sheets1(file_path)
    adjust_column_width(data, file_path)

    print(f"Processed {file_type} file: {file_path}")
    return file_path


def merge_files(folder_path, output_filename):
    print(f"กำลังรวมไฟล์จาก {folder_path}...")
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    if not files:
        flash(
            f'No files found in {folder_path}. Please upload some files first.', 'error')
        print(
            f"No files found in {folder_path}. Please upload some files first.")
        return False, f'No files found in {folder_path}. Please upload some files first.'

    merged_data = []
    for file in files:
        file_path = os.path.join(folder_path, file)
        print("กำลังอ่านไฟล์", os.path.basename(file_path))
        df = pd.read_excel(file_path, engine='openpyxl')
        # แปลงทุกคอลัมน์เป็น string แล้วลบช่องว่าง
        df = df.astype(str).apply(lambda x: x.str.strip())
        # df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        df = df.replace(r"[\t\n\r]", "", regex=True)
        df = df.reset_index(drop=True)

        if UPLOAD_FOLDER_FOOD in folder_path:
            columns_to_check = food_to_check
            num_cols = ["Unit Price", "Sales Count", "Sales"]
            df[num_cols] = df[num_cols].apply(lambda x: x.astype(
                str).str.replace(r"[\s,]", "", regex=True))
            print("Unit Price(ก่อนรวม)", df["Unit Price"].astype(
                str).apply(lambda x: f"[{x}]" if x.isspace() else x))
            print("Sales(ก่อนรวม)", df["Sales"].astype(str).apply(
                lambda x: f"[{x}]" if x.isspace() else x))
            print(df[num_cols])
            df[num_cols] = df[num_cols].apply(pd.to_numeric, errors="coerce")

        if UPLOAD_FOLDER_COUNTRY in folder_path:
            columns_to_check = country_to_check
            num_cols = ["Room", "Folio Number"]
            df[num_cols] = df[num_cols].apply(pd.to_numeric, errors="coerce")

        else:
            columns_to_check = None

        print(f"🔍 พบแถวที่ซ้ำกันทั้งหมด: {df.duplicated().sum()} แถว")

        merged_data.append(df)

    if len(merged_data) == 1:
        merged_df = merged_data[0]
    merged_df = pd.concat(merged_data, ignore_index=True)

    print("กำลังรวมไฟล์", output_filename)
    output_path = os.path.join(MERGE_FOLDER, output_filename)
    merged_df = remove_duplicate_rows(merged_df, columns_to_check)
    merged_df.to_excel(output_path, index=False)
    rename_sheet_to_sheets1(output_path)
    adjust_column_width(merged_df, output_path)
    return True, f'Successfully merged files into {output_filename}'

# ฟังก์ชันสำหรับรวมไฟล์จาก Food และ Country โดยใช้ Primary Keyss


def merge_files_by_primary_keys():
    print("กำลังตรวจสอบว่าไฟล์ Food และ Country มีหรือไม่...")
    food_file_path = os.path.join(MERGE_FOLDER, 'merged_food.xlsx')
    country_file_path = os.path.join(MERGE_FOLDER, 'merged_country.xlsx')

    if not os.path.exists(food_file_path) or not os.path.exists(country_file_path):
        print("ไม่พบไฟล์ Food หรือ Country หรือทั้งสองไฟล์")
        return None

    print("ไฟล์ทั้งสองมีอยู่ เริ่มการอ่านข้อมูล...")
    food_df = pd.read_excel(food_file_path)
    country_df = pd.read_excel(country_file_path)

    # ตรวจสอบ Primary Key ในทั้งสอง DataFrame
    print("กำลังแปลงคอลัมน์วันที่เป็น datetime...")
    try:
        food_df['Business Date'] = pd.to_datetime(food_df['Business Date'])
        country_df['Arrival Date'] = pd.to_datetime(country_df['Arrival Date'])
        country_df['Departure Date'] = pd.to_datetime(
            country_df['Departure Date'])
    except Exception as e:
        print(f"เกิดข้อผิดพลาดในการแปลงวันที่: {e}")
        return None

    # Merge ข้อมูลระหว่าง Food และ Country ตาม Primary Key
    print("กำลังทำการรวมข้อมูลระหว่าง Food และ Country ตาม Primary Key...")
    merged_df = pd.merge(
        country_df, food_df, left_on="Check Number", right_on="Check #", how="inner")

    # กรองข้อมูลโดยใช้เงื่อนไขวันที่
    merged_df = merged_df[
        (merged_df['Business Date'] >= merged_df['Arrival Date']) &
        (merged_df['Business Date'] <= merged_df['Departure Date'])
    ]

    # ตรวจสอบว่ามีข้อมูลหรือไม่หลังจากกรอง
    if merged_df.empty:
        print("ไม่พบข้อมูลที่ตรงกับเงื่อนไขที่กำหนด")
        return None

    # บันทึกผลลัพธ์ที่รวมกันแล้ว
    print("กำลังบันทึกผลลัพธ์ที่รวมกันแล้วลงไฟล์...")
    merged_output_path = os.path.join(MERGE_FOLDER, 'merged_output.xlsx')
    merged_df.to_excel(merged_output_path, index=False)

    print("รวมไฟล์และกรองข้อมูลสำเร็จแล้ว")
    rename_sheet_to_sheets1(merged_output_path)
    adjust_column_width(merged_df, merged_output_path)
    return merged_output_path


# ฟังก์ชันสำหรับการวิเคราะห์ข้อมูล
def analyze_data():
    merged_output_path = os.path.join(MERGE_FOLDER, 'merged_output.xlsx')

    print("ตรวจสอบว่าไฟล์ที่รวมแล้วมีอยู่หรือไม่...")
    if not os.path.exists(merged_output_path):
        flash("Merged file not found. Please merge the files first.", "error")
        print("ไม่พบไฟล์ที่รวมแล้ว! กรุณารวมไฟล์ก่อน")
        return None

    # อ่านข้อมูลจากไฟล์ที่รวมกันแล้ว
    print("กำลังอ่านข้อมูลจากไฟล์ที่รวมแล้ว...")
    merged_df = pd.read_excel(merged_output_path)

    #  กรองเฉพาะ Major Group ที่ต้องการเท่านั้น
    print("กำลังกรอง Major Group เฉพาะกลุ่มที่ต้องการ...")
    valid_major_groups = [
        "Food", "Wine", "Beer", "Non Alc Bev",
        "Spirit", "Misc", "Tobacco", "Gift Shop",
        "Boutique"
    ]
    merged_df = merged_df[merged_df["Major Group"].isin(valid_major_groups)]

    # วิเคราะห์ข้อมูล (คล้าย Pivot Table)
    print("กำลังทำการวิเคราะห์ข้อมูล...")
    analysis_df = merged_df.groupby(
        ['Business Date', 'Country', 'Major Group', 'Menu Item Name'],
        as_index=False
    )['Sales Count'].sum()

    # เรียงข้อมูลตามวันที่
    analysis_df = analysis_df.sort_values(by='Business Date')

    # บันทึกผลการวิเคราะห์เป็น Excel
    analysis_output_path = os.path.join(MERGE_FOLDER, 'analyzed_output.xlsx')
    print(f"กำลังบันทึกผลการวิเคราะห์ไปที่ {analysis_output_path} ...")
    analysis_df.to_excel(analysis_output_path, index=False)

    rename_sheet_to_sheets1(analysis_output_path)
    adjust_column_width(analysis_df, analysis_output_path)

    print("วิเคราะห์ข้อมูลเสร็จสมบูรณ์")
    flash("Data analysis completed successfully.", "success")
    return analysis_output_path


@app.route("/upload", methods=["POST"])
def upload_files():
    food_file_path = None
    country_file_path = None

    food_file = request.files.get('foodFile')
    country_file = request.files.get('countryFile')

    if not food_file and not country_file:
        print("ไม่พบไฟล์")
        return jsonify({"status": "error", "message": "No files uploaded"}), 400

    if food_file and allowed_file(food_file.filename):
        print("พบไฟล์ food")
        food_file_path = process_uploaded_file(
            food_file, app.config['UPLOAD_FOLDER_FOOD'], "food")

    if country_file and allowed_file(country_file.filename):
        print("พบไฟล์ country")
        country_file_path = process_uploaded_file(
            country_file, app.config['UPLOAD_FOLDER_COUNTRY'], "country")

    if not food_file_path and not country_file_path:
        return jsonify({"status": "error", "message": "Invalid file(s)"}), 400

    return jsonify({
        "status": "success",
        "message": "Files processed successfully",
        "food_file": os.path.basename(food_file_path) if food_file_path else None,
        "country_file": os.path.basename(country_file_path) if country_file_path else None
    })


@app.route("/merge", methods=["POST"])
def merge():
    data = request.get_json()
    merge_option = data.get('mergeOption')
    filecountry = "merged_country.xlsx"
    filefood = "merged_food.xlsx"
    if merge_option == "mergeFood":
        success, message = merge_files(
            app.config['UPLOAD_FOLDER_FOOD'], filefood)
    elif merge_option == "mergeCountry":
        success, message = merge_files(
            app.config['UPLOAD_FOLDER_COUNTRY'], filecountry)
    elif merge_option == "mergeBoth":
        success, message = merge_files(
            app.config['UPLOAD_FOLDER_FOOD'], filefood)
        if success:
            success, message = merge_files(
                app.config['UPLOAD_FOLDER_COUNTRY'], filecountry)
    else:
        success, message = False, "Invalid option."
    return jsonify({'success': success, 'message': message})


@app.route("/process", methods=["POST"])
def process_action():
    data = request.get_json()
    process_option = data.get('processOption')
    if process_option == 'merge':
        output_path = merge_files_by_primary_keys()
        if output_path:
            return jsonify({'success': True, 'message': f"File merged successfully: {output_path}"})
        else:
            return jsonify({'success': False, 'message': "Error merging files or no data."})
    elif process_option == 'analyze':
        output_path = analyze_data()
        if output_path:
            return jsonify({'success': True, 'message': f"Analysis completed. File saved at: {output_path}"})
        else:
            return jsonify({'success': False, 'message': "Error analyzing data or merged file not found."})
    else:
        return jsonify({'success': False, 'message': "Invalid process option."})


@app.route('/get_files', methods=['POST'])
def get_files():
    folder = request.json.get('folder')
    if folder == 'Food':
        folder_path = UPLOAD_FOLDER_FOOD
    elif folder == 'Country':
        folder_path = UPLOAD_FOLDER_COUNTRY
    else:
        return jsonify(files=[])

    if not os.path.exists(folder_path):
        return jsonify(files=[])

    files = os.listdir(folder_path)
    return jsonify(files=files)


@app.route('/delete_file', methods=['POST'])
def delete_file():
    data = request.json
    folder = data.get('folder')
    filename = data.get('filename')

    folder_path = None
    if folder == 'Food':
        folder_path = UPLOAD_FOLDER_FOOD
    elif folder == 'Country':
        folder_path = UPLOAD_FOLDER_COUNTRY

    if folder_path is None:
        return jsonify(success=False, message="Invalid folder")

    file_path = os.path.join(folder_path, filename)

    if os.path.exists(file_path):
        os.remove(file_path)
        return jsonify(success=True, message=f"Deleted {filename}")
    return jsonify(success=False, message="File not found")


@app.route("/get_data", methods=["GET"])
def get_data():
    try:
        # รับค่าจาก Query Parameters
        country = request.args.get("country", "").strip()
        top_rank = request.args.get("toprank", "").strip()
        food_type = request.args.get("foodType", "").strip()
        date_from_str = request.args.get("dateFrom", "").strip()
        date_to_str = request.args.get("dateTo", "").strip()

        if country == "United Kingdom":
            country = "Great Britain"

        if not country or not date_from_str or not date_to_str:
            return jsonify({"error": "Missing required parameters"}), 400

        try:
            top_rank = int(top_rank) if top_rank.isdigit() else 100

        except ValueError:
            return jsonify({"error": "Invalid value for topRank"}), 400

        # แปลงวันที่
        date_from = pd.to_datetime(
            date_from_str, format="%d/%m/%Y", errors="coerce")
        date_to = pd.to_datetime(
            date_to_str, format="%d/%m/%Y", errors="coerce")

        if pd.isna(date_from) or pd.isna(date_to):
            return jsonify({"error": "Invalid date format. Use DD/MM/YYYY"}), 400

        if date_from > date_to:
            return jsonify({"error": "dateFrom must be earlier than or equal to dateTo"}), 400

        # กรองข้อมูลตามวันที่และประเทศ
        filtered_df = df[(df["Business Date"] >= date_from) & (
            df["Business Date"] <= date_to) & (df["Country"] == country)]

        # กรณี foodType ไม่ใช่ "All" ให้กรองเฉพาะประเภทนั้น
        if food_type and food_type != "All":
            # กรณีที่กรองโดย Major Group (food_type != "All")
            filtered_df = filtered_df[filtered_df["Major Group"] == food_type]

            # ถ้าหลังจากกรองแล้วไม่มีข้อมูล ให้คืน []
            if filtered_df.empty:
                return jsonify([])

            # รวม Count ตาม "Major Group" และ "Menu Item Name"
            grouped_df = (filtered_df.groupby(["Major Group", "Menu Item Name"])["Sales Count"]
                          .sum()
                          .reset_index()
                          .sort_values(by="Sales Count", ascending=False))
            grouped_df["Sales Count"] = grouped_df["Sales Count"].round(
                0).astype(int)  # เลขจำนวนเต็ม

            # แสดง 10 อันดับแรก หรือเท่าที่มี
            top_items = grouped_df.head(min(top_rank, len(grouped_df)))

            # จัดโครงสร้าง JSON ให้เป็นแบบที่ต้องการ
            result = [{
                "Major Group": food_type,
                "Menu Item Name": top_items.to_dict(orient="records")
            }]

        else:
            # กรณี foodType = "All" ให้แสดง 10 อันดับแรกของแต่ละ foodType
            result = []
            for food_type, food_type_df in filtered_df.groupby("Major Group"):
                top_items = (
                    food_type_df.groupby("Menu Item Name")["Sales Count"]
                    .sum()
                    .reset_index()
                    .sort_values(by="Sales Count", ascending=False)
                )

                # แปลง Sales Count เป็นจำนวนเต็ม
                top_items["Sales Count"] = top_items["Sales Count"].round(
                    0).astype(int)

                # เลือก 10 อันดับแรกหรือตามจำนวนที่มี
                top_items = top_items.head(min(top_rank, len(top_items)))

                result.append({
                    "foodType": food_type,
                    "Menu Item Name": top_items.to_dict(orient="records")
                })

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
